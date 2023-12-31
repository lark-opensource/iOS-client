import openpyxl

# Open the Excel file
workbook = openpyxl.load_workbook('example.xlsx')

# Select the active sheet
sheet = workbook.active

# Loop through each row in the sheet and generate Swift code
swift_code = ''

#将带“-”的颜色格式转换为驼峰命名式的
def convert_to_camel_case(string):
    # Split the string into components using the hyphen as a separator
    components = string.split("-")

    # Convert the first component to lowercase and the rest to title case
    camel_case_components = [components[0]] + [c.title() for c in components[1:]]

    # Join the components and return the resulting camel case string
    return "".join(camel_case_components)

#生成如下格式代码：
#        '/// rgb value: {}\n'.format(bg_color)
#'public static var {}: UIColor {{ return UDColor.{} }}\n\n'.format(color_name, color_name)
#或 'public static var {}: UIColor {{ return UDColor.{} }}\n\n'.format(color_name, color_name)
def rewriteBaseColor():
    global swift_code
    for row in sheet.iter_rows():
        color_name = row[0].value
        bg_color = row[1].value
        text_color = row[2].value
        bg_color_hex = bg_color[0:]  # Remove the # character from the hex code
        text_color_hex = text_color[0:]

        swift_code += '/// rgb value: {}\n'.format(bg_color)
        # swift_code += 'public static var {}: UIColor {{ return rgb({}) & rgb({}) }}\n\n'.format(color_name, bg_color_hex, text_color_hex)
        swift_code += 'public static var {}: UIColor {{ return UDColor.{} }}\n\n'.format(color_name, color_name)

#生成如下格式代码：
#        '/// {}, Value: "{}"\n'.format(camelCaseStr, id)
#'public static var {}: UIColor {{ return rgb({}) & rgb({}) }}\n\n'.format(color_name, bg_color_hex, text_color_hex) 或
#'static let {} = UDColor.Name("{}")\n\n'.format(camelCaseStr, id)
def rewriteToken():
    global swift_code
    for row in sheet.iter_rows():
        id = row[1].value
        camelCaseStr = convert_to_camel_case(id)
        # /// primaryPri50, Value: "primary-pri-50" .
        swift_code += '/// {}, Value: "{}"\n'.format(camelCaseStr, id)
        # swift_code += 'public static var {}: UIColor {{ return rgb({}) & rgb({}) }}\n\n'.format(color_name, bg_color_hex, text_color_hex)
        swift_code += 'static let {} = UDColor.Name("{}")\n\n'.format(camelCaseStr, id)


#生成'store[UDColor.Name("{}")] = {}\n'.format(id, color)格式代码
def replaceDmLm():
    global swift_code
    for row in sheet.iter_rows():
        id = row[1].value
        lightColor = row[6].value
        darkColor = row[8].value
        if lightColor == darkColor:
            color = transformColor(lightColor)
        else:
            color = '{} & {}'.format(transformColor(lightColor), transformColor(darkColor))
        swift_code += 'store[UDColor.Name("{}")] = {}\n'.format(id, color)



#将color的格式从xlsx文件中提供的格式转换为UDColor
def transformColor(colors: str):
    alpha = ''
    if colors.__contains__(","): 
        color, alpha = colors.split(",")
        alpha = '.withAlphaComponent({})'.format(str(int(alpha[:-1]) / 100.0))
    else:
        color = colors
    result = ''
    # #23983f
    if color.startswith("#"):
        result += 'UDColor.rgb({})'.format(color.replace("#", "0x"))
    # N100
    elif color.startswith(color[0]) and color[0].isupper():
        result += 'UDColor.{}'.format(color)
    # primary/pri-400
    else:
        result += 'UDColor.'
        result += convert_to_camel_case(color.replace('/', '-'))
    result += alpha
    return result

#生成如下格式代码：
#        '/// Light: {}, Dark: {}\n'.format(transformColor(lightColor), transformColor(darkColor))
#        'public static var {}: UIColor {{\n'.format(convert_to_camel_case(id))
#        '    return UDColor.getValueByKey(.{}) ?? {}\n'.format(convert_to_camel_case(id), color)
#        '}\n\n'
def generateUDColor():
    global swift_code
    for row in sheet.iter_rows():
        color = ''
        id = row[1].value
        lightColor = row[6].value
        darkColor = row[8].value
        if lightColor == darkColor:
            color = transformColor(lightColor)
        else:
            color = '{} & {}'.format(transformColor(lightColor), transformColor(darkColor))
        swift_code += '/// Light: {}, Dark: {}\n'.format(transformColor(lightColor), transformColor(darkColor))
        swift_code += 'public static var {}: UIColor {{\n'.format(convert_to_camel_case(id))
        swift_code += '    return UDColor.getValueByKey(.{}) ?? {}\n'.format(convert_to_camel_case(id), color)
        swift_code += '}\n\n'


    # 生成如下格式代码
    # /// primaryPri50, Value: UDColor.B50 & UDColor.B50
    # public static var primaryPri50: UIColor { return UDColor.primaryPri50 }
def generate(col1, col2):
    global swift_code
    for row in sheet.iter_rows():
        color = ''
        id = row[1].value
        lightColor = row[col1].value
        darkColor = row[col2].value
        if lightColor == darkColor:
            color = transformColor(lightColor)
        else:
            color = '{} & {}'.format(transformColor(lightColor), transformColor(darkColor))
        swift_code += '/// {}, Value: {} & {}\n'.format(convert_to_camel_case(id), transformColor(lightColor), transformColor(darkColor))
        swift_code += '@objc public static var {}: UIColor {{ return UDColor.{} }}\n\n'.format(convert_to_camel_case(id), convert_to_camel_case(id))

#将整个库中的颜色按example.xlsx中按第一列到第二列替换
def generate_script():
    global swift_code
    for row in sheet.iter_rows():
        color1 = row[0].value
        color2 = row[1].value
        if color1 is not None:
            color1 = convert_to_camel_case(color1.replace('/', '-'))
        if color2 is not None:
            color2 = convert_to_camel_case(color2.replace('/', '-'))
        swift_code += 'new_content = content.replace("{}", "{}")\n'.format(color1, color2)



generate_script()
# Write the Swift code to a file
with open('script.swift', 'w') as file:
    file.write(swift_code)
